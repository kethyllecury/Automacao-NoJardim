import os
import time
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager

load_dotenv("cripto.env")

email = os.getenv("emailtakeat")
password = os.getenv("senhatakeat")

hoje = datetime.today()
data_atual = hoje.strftime("%d-%m-%Y")
data = datetime.now()

mes_atual = f"{data.month:02}"
mes_atual_int = datetime.now().month

ano_atual = datetime.now().year

caminho = r'G:\\.shortcut-targets-by-id\\1ySpnxdv_XzDx42T-TJ1JYAs013TNGwLO\\SigaBPO - Drive Arquivos\\No Jardiim\\Demonstrativos\\Base Relatório\\faturamento_atualizado.xlsx'

def configurar_driver():
    edge_options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": r"C:\Users\sigab\Downloads",
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    edge_options.add_experimental_option("prefs", prefs)
    edge_options.add_argument("--start-maximized")
    
    driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()), options=edge_options)
    return driver

def realizar_login(driver):
    driver.get("https://gestor.takeat.app/login")
    
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="E-mail"]')))
    email_field = driver.find_element(By.XPATH, '//input[@placeholder="E-mail"]')
    email_field.send_keys(email)

    senha_field = driver.find_element(By.XPATH, '//input[@placeholder="Senha"]')
    senha_field.send_keys(password)

    button_acesso = driver.find_element(By.XPATH, '//button[@type="button"]//span[text()="Acessar"]')
    button_acesso.click()

def remover_anuncios(driver, max_tentativas=3):
    
    wait = WebDriverWait(driver, 10)  
    tentativas = 0

    while tentativas < max_tentativas:
        fechado = False  

        try:
            botao_copiar = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Copiar link para compartilhar!')]")))
            botao_copiar.click()
            print("Anúncio fechado (botão copiar)!")
            fechado = True
        except Exception as e:
            print("Nenhum anúncio (botão copiar).", str(e))

        try:
            botao_fechar = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@style='padding: 10px; align-self: flex-end; cursor: pointer;']")))
            botao_fechar.click()
            print("Anúncio fechado (botão fechar)!")
            fechado = True
        except Exception as e:
            print("Nenhum anúncio (botão fechar).", str(e))

        if not fechado:
            print("encerrando verificação.")
            break
        
        tentativas += 1
        time.sleep(2)

def selecionar_relatorio(driver):

    wait = WebDriverWait(driver, 20)
    relatorio_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[span[text()='Relatórios']]")))
    relatorio_link.click()

    wait = WebDriverWait(driver, 20)
    elemento = wait.until(EC.visibility_of_element_located((By.ID, "menu-item-1-2")))
    elemento.click()

def verificar_fim_de_semana(data_atual):
    data_obj = datetime.strptime(data_atual, "%d-%m-%Y")
    print(f"Data fornecida: {data_atual}, Data convertida: {data_obj}")

    data_anterior = (data_obj - timedelta(days=1)).strftime("%d")
    print(f"Data anterior: {data_anterior}")
    
    if data_obj.weekday() == 0: 
        sexta = (data_obj - timedelta(days=3)).strftime("%d")
        sabado = (data_obj - timedelta(days=2)).strftime("%d")
        print(f"Segunda-feira detectada, retornando sexta e sábado: {sexta}, {sabado}")
        return [sexta, sabado, data_anterior]
    else:
        return [data_anterior]

def gerar_relatorio(driver, datas):
        
        time.sleep(40)
    
        wait = WebDriverWait(driver, 30) 
        elementos = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-jQAyio")))
        elementos[0].click()

        dates = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-GvgMv")))
        date_24 = [el for el in dates if el.text == datas]
        if len(date_24) > 1:
            date_24[1].click()

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[0].click()

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[1].click()

        dates = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-GvgMv")))
        date_24 = [el for el in dates if el.text == datas]
        if len(date_24) > 1:
            date_24[1].click()

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[1].click()

        baixar = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-fujznN")))
        baixar.click()
        time.sleep(10)

def dividir_mesa_data_hora(valor):

    partes = valor.split(' - ')  
    if len(partes) == 3:
        return partes  
    else:
        return [None, None, None] 

def tratar_planilha(datas, mes_atual):

    global arquivo_saida
    global planilha
    global df

    planilha = fr"C:\Users\sigab\Downloads\Faturamento({datas}-{mes_atual}_{datas}-{mes_atual}).xlsx"

    df = pd.read_excel(planilha, sheet_name="Relatório Produtos")

    df.insert(0, "Modalidade", "")
    df.insert(1,"Competencia","")
    df.insert(2,"Dias","")
    df.insert(3,"Tipo Pgto","")
    df.insert(4,"Local","")
    df.insert(5,"Data","")
    df.insert(6,"Hora","")
    df.insert(14,"qtd", 1)
    df.insert(15,"Pix","" )
    df.insert(16,"PIX","" )


    valores_desejados = [
    "Pix", "Crédito", "Débito", "Dinheiro", "Clube", "iFood", "Resgate Clube",
    "Visa Crédito", "Elo Débito", "MasterCard Crédito", "MasterCard Débito",
    "Visa Débito", "Elo Crédito", "Pagamento Online iFood"
    ]

    df.loc[df.iloc[:, 7].isin(valores_desejados), df.columns[3]] = df.iloc[:, 7]

    df.iloc[:, 0] = df.iloc[:, 0].replace('', pd.NA)
    df.iloc[:, 0] = df.iloc[:, 0].ffill()

    df.iloc[:, 4] = df.iloc[:, 4].replace('Balcão -1', 'Balcão 1')

    df.iloc[:, 4], df.iloc[:, 5], df.iloc[:, 6] = zip(*df.iloc[:, 7].apply(dividir_mesa_data_hora))

    df = df.drop(df.columns[7], axis=1)
    df = df.drop(0,axis=0)
    df = df.dropna(subset=[df.columns[4]])
    df['Data'] = pd.to_datetime(df['Data'], errors='coerce', dayfirst=True)
    print(df.head(10))  

    arquivo_saida = planilha
    df.to_excel(arquivo_saida, index=False, engine='openpyxl')

def localizar_ultima_linha(arquivo_saida, caminho):
    
    global planilha1
    global planilha2
    global ultima_linha_df

    planilha1 = pd.read_excel(arquivo_saida)
    planilha2 = pd.read_excel(caminho)

    print("Colunas da planilha 1:", planilha1.columns)
    print("Colunas da planilha 2:", planilha2.columns)

    ultima_linha_df = planilha2['Modalidade'].last_valid_index() + 2
    print(f"Última linha válida na aba 'produtos': {ultima_linha_df}")

def gerar_arquivo(ultima_linha_df, planilha1, planilha2):

    df_concatenado = pd.concat([planilha2.iloc[: ultima_linha_df], planilha1], ignore_index=True)

    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df_concatenado.to_excel(writer, index=False, sheet_name="Relatório Venda")

    wb = load_workbook(caminho, data_only=False)
    ws = wb["Relatório Venda"] 

    for i, row in df_concatenado.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)

    time.sleep(60)

    for row in range(2, len(df_concatenado) + 2):
        ws[f'A{row}'] = f'=VLOOKUP(D{row}, O:P, 2, 0)'
        ws[f'B{row}'] = f'=MONTH(F{row})&YEAR(F{row})'
        ws[f'C{row}'] = f'=IF(F{row}=F{row-1}, 0, 1)' 
        
    
    time.sleep(5)

    wb.save(caminho)

def deletar_arquivo(planilha):
    if os.path.exists(planilha):
        os.remove(planilha)
        print(f"Arquivo '{planilha}' deletado com sucesso.")
    else:
        print("Nenhum dos arquivos foi encontrado para deletar.")

driver = configurar_driver()
realizar_login(driver)
remover_anuncios(driver)
selecionar_relatorio(driver)
datas_para_processar = verificar_fim_de_semana(data_atual)
for datas in datas_para_processar:
    print(f"Iniciando o processamento para a data: {datas}")
    gerar_relatorio(driver, datas)
    tratar_planilha(datas, mes_atual)
    localizar_ultima_linha(arquivo_saida, caminho)
    gerar_arquivo(ultima_linha_df, planilha1, planilha2)
    deletar_arquivo(planilha)

print("Processo concluído.")
driver.quit()

