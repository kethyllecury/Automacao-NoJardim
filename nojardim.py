import os
import time
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager

load_dotenv("cripto.env")

email = os.getenv("emailtakeat")
password = os.getenv("senhatakeat")

hoje = datetime.today()
data_atual = hoje.strftime("%d-%m-%Y")
data = datetime.now()

mes_atual = f"{data.month:02}"
mes_atual_int = datetime.now().month

ano_atual = datetime.now().year

caminho = f'G:\\.shortcut-targets-by-id\\1ySpnxdv_XzDx42T-TJ1JYAs013TNGwLO\\SigaBPO - Drive Arquivos\\No Jardiim\\Demonstrativos\\Base Relatório\\produtos_atualizados.xlsx'

def configurar_driver():
    edge_options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": r"C:\Users\sigab\Downloads",
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    edge_options.add_experimental_option("prefs", prefs)
    edge_options.add_argument("--start-maximized")
    
    driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()), options=edge_options)
    return driver

def realizar_login(driver):
    driver.get("https://gestor.takeat.app/login")
    
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="E-mail"]')))
    email_field = driver.find_element(By.XPATH, '//input[@placeholder="E-mail"]')
    email_field.send_keys(email)

    senha_field = driver.find_element(By.XPATH, '//input[@placeholder="Senha"]')
    senha_field.send_keys(password)

    button_acesso = driver.find_element(By.XPATH, '//button[@type="button"]//span[text()="Acessar"]')
    button_acesso.click()

def remover_anuncios(driver, max_tentativas=3):
    
    wait = WebDriverWait(driver, 10)  
    tentativas = 0

    while tentativas < max_tentativas:
        fechado = False  

        try:
            botao_copiar = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Copiar link para compartilhar!')]")))
            botao_copiar.click()
            print("Anúncio fechado (botão copiar)!")
            fechado = True
        except Exception as e:
            print("Nenhum anúncio (botão copiar).", str(e))

        try:
            botao_fechar = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@style='padding: 10px; align-self: flex-end; cursor: pointer;']")))
            botao_fechar.click()
            print("Anúncio fechado (botão fechar)!")
            fechado = True
        except Exception as e:
            print("Nenhum anúncio (botão fechar).", str(e))

        if not fechado:
            print("encerrando verificação.")
            break
        
        tentativas += 1
        time.sleep(2)

def selecionar_relatorio(driver):

    wait = WebDriverWait(driver, 20)
    relatorio_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[span[text()='Relatórios']]")))
    relatorio_link.click()

    wait = WebDriverWait(driver, 20)
    elemento = wait.until(EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Vendidos')]")))
    elemento.click()

def verificar_fim_de_semana(data_atual):
    data_obj = datetime.strptime(data_atual, "%d-%m-%Y")
    print(f"Data fornecida: {data_atual}, Data convertida: {data_obj}")

    data_anterior = (data_obj - timedelta(days=1)).strftime("%d")
    print(f"Data anterior: {data_anterior}")
    
    if data_obj.weekday() == 0: 
        sexta = (data_obj - timedelta(days=3)).strftime("%d")
        sabado = (data_obj - timedelta(days=2)).strftime("%d")
        print(f"Segunda-feira detectada, retornando sexta e sábado: {sexta}, {sabado}, {data_anterior}")
        return [sexta, sabado, data_anterior]
    else:
        return [data_anterior]

def gerar_relatorio(driver, datas):
        
        time.sleep(40)
    
        wait = WebDriverWait(driver, 30) 
        elementos = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-jQAyio")))
        elementos[0].click()

        dates = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-GvgMv")))
        date_24 = [el for el in dates if el.text == datas]
        if len(date_24) > 1:
            date_24[1].click()  

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[0].click()

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[1].click()

        dates = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "sc-GvgMv")))
        date_24 = [el for el in dates if el.text == datas]
        if len(date_24) > 1:
            date_24[1].click()

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-jQAyio")))
        elementos[1].click()

        buscar = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Buscar']]")))
        buscar.click()
        print("Botão 'Buscar' clicado com sucesso!")
        
        time.sleep(10)

        baixar = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sc-fujznN")))
        baixar.click()
        time.sleep(10)

def tratar_planilha(datas, mes_atual):

    global arquivo_saida
    global planilha

    planilha = fr"C:\Users\sigab\Downloads\Produtos({datas}-{mes_atual}_{datas}-{mes_atual}).xlsx"

    df = pd.read_excel(planilha, sheet_name="Relatório Produtos")

    dia = int(datas)

    data_formatada = datetime(hoje.year, hoje.month, dia)

    df.insert(0, "Período", data_formatada)
    df.insert(1,"Tipo","")
    df.insert(len(df.columns), "PIZZA de Dois Sabores", "")

    arquivo_saida = planilha
    df.to_excel(arquivo_saida, index=False, engine='openpyxl')

def renomear_planilhas(arquivo_saida, caminho):

    global planilha1
    global planilha2
    global ultima_linha_df

    planilha1 = pd.read_excel(arquivo_saida)
    planilha2 = pd.read_excel(caminho)

    planilha2.rename(columns={'Tipo 1': 'Tipo'}, inplace=True)
    

    print("Colunas da planilha 1:", planilha1.columns)
    print("Colunas da planilha 2:", planilha2.columns)

    ultima_linha_df = planilha2['Período'].last_valid_index() + 2
    print(f"Última linha válida na aba 'produtos': {ultima_linha_df}")

def gerar_arquivo(ultima_linha_df, planilha1, planilha2):

    df_concatenado = pd.concat([planilha2.iloc[: ultima_linha_df], planilha1], ignore_index=True)

    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df_concatenado.to_excel(writer, index=False, sheet_name="Relatório Produtos")

    wb = load_workbook(caminho, data_only=False)
    ws = wb["Relatório Produtos"] 

    for i, row in df_concatenado.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)

    time.sleep(60)

    for row in range(2, len(df_concatenado) + 2):
        ws[f'B{row}'] = f'=IF(C{row}=$N$1,"Normal",IF(LEFT(D{row},4)="  - ","Complemento","Normal"))'

    time.sleep(5)

    wb.save(caminho)

def deletar_arquivo(planilha):
    if os.path.exists(planilha):
        os.remove(planilha)
        print(f"Arquivo '{planilha}' deletado com sucesso.")
    else:
        print("Nenhum dos arquivos foi encontrado para deletar.")

driver = configurar_driver()
realizar_login(driver)
remover_anuncios(driver)
selecionar_relatorio(driver)
datas_para_processar = verificar_fim_de_semana(data_atual)
for datas in datas_para_processar:
    print(f"Iniciando o processamento para a data: {datas}")
    gerar_relatorio(driver, datas)
    tratar_planilha(datas, mes_atual)
    renomear_planilhas(arquivo_saida, caminho)
    gerar_arquivo(ultima_linha_df, planilha1, planilha2)
    deletar_arquivo(planilha)

print("Processo concluído.")
driver.quit()

